from io import BytesIO, StringIO
import csv
import logging
from typing import List
from urllib.parse import urlparse
from uuid import UUID

import httpx
from fastapi import UploadFile
from langfuse import get_client
from openpyxl import load_workbook

from src.dataset.models import (
    DocumentORM,
    QuestionORM,
)
from src.dataset.repository import (
    find_document_by_name,
    get_dataset,
    get_documents_by_dataset,
    get_questions_by_dataset,
    insert_document,
    insert_question,
)
from src.infrastructure.blob_storage.blob import insert_blob

logger = logging.getLogger(__name__)

_XLSX_COLUMNS = ("query", "answer", "filename")
_DOCUMENTS_XLSX_COLUMNS = ("file_name", "file_url")


class DocumentNotFoundError(LookupError):
    def __init__(self, row_index: int, name: str) -> None:
        super().__init__(
            f"Row {row_index}: no document matches filename={name!r}"
        )
        self.row_index = row_index
        self.name = name


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_xlsx_rows(data: bytes) -> list[tuple[str, str, str]]:
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read questions XLSX: {e}") from e
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []

        headers = [_normalize_header(c) for c in header_row]
        col_index = {h: i for i, h in enumerate(headers) if h}
        for name in _XLSX_COLUMNS:
            if name not in col_index:
                raise ValueError(
                    f"XLSX must contain columns {list(_XLSX_COLUMNS)}; missing {name!r}"
                )

        def cols(r: tuple) -> tuple[str, str, str]:
            def at(name: str) -> str:
                i = col_index[name]
                return _cell_str(r[i] if i < len(r) else None)

            return at("query"), at("answer"), at("filename")

        out: list[tuple[str, str, str]] = []
        for row in rows:
            if row is None:
                continue
            q, a, fn = cols(row)
            if not (q or a or fn):
                continue
            out.append((q, a, fn))
        return out
    finally:
        wb.close()


def _parse_csv_rows(data: bytes) -> list[tuple[str, str, str]]:
    # Use utf-8 with BOM tolerance; replace errors so ingestion can still validate rows.
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(StringIO(text))

    if reader.fieldnames is None:
        return []

    field_map: dict[str, str] = {}
    for raw in reader.fieldnames:
        norm = _normalize_header(raw)
        if norm:
            # if duplicates exist after normalization, keep the first one
            field_map.setdefault(norm, raw)

    for name in _XLSX_COLUMNS:
        if name not in field_map:
            raise ValueError(
                f"CSV must contain columns {list(_XLSX_COLUMNS)}; missing {name!r}"
            )

    out: list[tuple[str, str, str]] = []
    for row in reader:
        q = _cell_str(row.get(field_map["query"]))
        a = _cell_str(row.get(field_map["answer"]))
        fn = _cell_str(row.get(field_map["filename"]))
        if not (q or a or fn):
            continue
        out.append((q, a, fn))
    return out


def _normalize_header_key(value: object) -> str:
    """
    Header normalization that tolerates 'file name' / 'file-name' / 'File_Name' etc.
    """
    return _normalize_header(value).replace(" ", "_").replace("-", "_")


def _parse_documents_xlsx_rows(data: bytes) -> list[tuple[str, str]]:
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read documents XLSX: {e}") from e
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []

        headers = [_normalize_header_key(c) for c in header_row]
        col_index = {h: i for i, h in enumerate(headers) if h}

        # Back-compat: accept `filename` as alias for `file_name`
        if "file_name" not in col_index and "filename" in col_index:
            col_index["file_name"] = col_index["filename"]

        for name in _DOCUMENTS_XLSX_COLUMNS:
            if name not in col_index:
                raise ValueError(
                    f"XLSX must contain columns {list(_DOCUMENTS_XLSX_COLUMNS)}; missing {name!r}"
                )

        out: list[tuple[str, str]] = []
        for row in rows:
            if row is None:
                continue
            file_name = _cell_str(
                row[col_index["file_name"]]
                if col_index["file_name"] < len(row)
                else None
            )
            file_url = _cell_str(
                row[col_index["file_url"]] if col_index["file_url"] < len(row) else None
            )
            if not (file_name or file_url):
                continue
            out.append((file_name, file_url))
        return out
    finally:
        wb.close()


def _parse_documents_csv_rows(data: bytes) -> list[tuple[str, str]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(StringIO(text))
    if reader.fieldnames is None:
        return []

    field_map: dict[str, str] = {}
    for raw in reader.fieldnames:
        norm = _normalize_header_key(raw)
        if norm:
            field_map.setdefault(norm, raw)

    # Back-compat: accept `filename` as alias for `file_name`
    if "file_name" not in field_map and "filename" in field_map:
        field_map["file_name"] = field_map["filename"]

    for name in _DOCUMENTS_XLSX_COLUMNS:
        if name not in field_map:
            raise ValueError(
                f"CSV must contain columns {list(_DOCUMENTS_XLSX_COLUMNS)}; missing {name!r}"
            )

    out: list[tuple[str, str]] = []
    for row in reader:
        file_name = _cell_str(row.get(field_map["file_name"]))
        file_url = _cell_str(row.get(field_map["file_url"]))
        if not (file_name or file_url):
            continue
        out.append((file_name, file_url))
    return out


def _url_host(url: str) -> str:
    """Log-friendly host only (no path/query; avoids leaking SAS tokens)."""
    try:
        return urlparse(url).netloc or "unknown"
    except Exception:
        return "invalid"


def _file_ext(filename: str | None) -> str:
    if not filename:
        return ""
    parts = filename.rsplit(".", 1)
    return parts[-1].lower() if len(parts) == 2 else ""


def _parse_questions_rows(*, raw: bytes, filename: str | None) -> list[tuple[str, str, str]]:
    ext = _file_ext(filename)
    if ext == "csv":
        return _parse_csv_rows(raw)
    if ext == "xlsx":
        return _parse_xlsx_rows(raw)
    raise ValueError("Only .xlsx or .csv files are supported for questions ingestion")


def _parse_documents_rows(*, raw: bytes, filename: str | None) -> list[tuple[str, str]]:
    ext = _file_ext(filename)
    if ext == "csv":
        return _parse_documents_csv_rows(raw)
    if ext == "xlsx":
        return _parse_documents_xlsx_rows(raw)
    raise ValueError("Only .xlsx or .csv files are supported for documents ingestion")


def _upload_file_from_bytes(*, filename: str, content: bytes) -> UploadFile:
    # UploadFile wraps a sync file object; BytesIO is fine here.
    return UploadFile(filename=filename, file=BytesIO(content))


async def ingest_documents_from_file(
    file: UploadFile, dataset_id: UUID
) -> list[DocumentORM]:
    dataset = await get_dataset(dataset_id)
    if dataset is None:
        raise ValueError(f"Dataset {dataset_id} not found")

    try:
        raw = await file.read()
    except Exception as e:
        logger.exception(
            "dataset.ingest_documents.read_failed",
            extra={
                "event": "dataset.ingest_documents.read_failed",
                "dataset_id": str(dataset_id),
            },
        )
        raise ValueError(f"Failed to read upload file: {e}") from e

    try:
        rows = _parse_documents_rows(raw=raw, filename=file.filename)
    except ValueError:
        raise
    except Exception as e:
        logger.exception(
            "dataset.ingest_documents.parse_failed",
            extra={
                "event": "dataset.ingest_documents.parse_failed",
                "dataset_id": str(dataset_id),
            },
        )
        raise ValueError(f"Failed to parse upload file: {e}") from e
    logger.debug(
        "dataset.ingest_documents.parsed",
        extra={
            "event": "dataset.ingest_documents.parsed",
            "dataset_id": str(dataset_id),
            "row_count": len(rows),
            "upload_filename": file.filename,
        },
    )

    docs: list[DocumentORM] = []
    async with httpx.AsyncClient(follow_redirects=True, timeout=60) as client:
        for i, (file_name, file_url) in enumerate(rows, start=2):
            if not file_name:
                raise ValueError(f"Row {i}: file_name is required")
            if not file_url:
                raise ValueError(f"Row {i}: file_url is required")

            try:
                resp = await client.get(file_url)
                resp.raise_for_status()
            except httpx.HTTPError as e:
                logger.warning(
                    "dataset.ingest_documents.download_failed",
                    extra={
                        "event": "dataset.ingest_documents.download_failed",
                        "dataset_id": str(dataset_id),
                        "row_index": i,
                        "url_host": _url_host(file_url),
                        "error_type": type(e).__name__,
                        "error": str(e),
                    },
                )
                raise ValueError(
                    f"Row {i}: failed to download {file_url!r}: {e}"
                ) from e

            stored_name = f"{dataset.name}_{file_name}"
            upload = _upload_file_from_bytes(filename=stored_name, content=resp.content)
            try:
                doc = await ingest_document(
                    file=upload, file_url=file_url, dataset_id=dataset_id
                )
            except ValueError:
                raise
            except Exception as e:
                logger.exception(
                    "dataset.ingest_documents.row_store_failed",
                    extra={
                        "event": "dataset.ingest_documents.row_store_failed",
                        "dataset_id": str(dataset_id),
                        "row_index": i,
                        "url_host": _url_host(file_url),
                    },
                )
                raise ValueError(
                    f"Row {i}: failed to store document: {e}"
                ) from e
            docs.append(doc)
            logger.debug(
                "dataset.ingest_documents.row_stored",
                extra={
                    "event": "dataset.ingest_documents.row_stored",
                    "dataset_id": str(dataset_id),
                    "row_index": i,
                    "document_id": str(doc.id),
                    "url_host": _url_host(file_url),
                },
            )

    logger.info(
        "dataset.ingest_documents.complete",
        extra={
            "event": "dataset.ingest_documents.complete",
            "dataset_id": str(dataset_id),
            "inserted": len(docs),
        },
    )
    return docs


# Backwards-compatible name (router may still import it in older code).
ingest_documents_from_xlsx = ingest_documents_from_file


async def get_documents(dataset_id: UUID) -> list[DocumentORM]:
    dataset = await get_dataset(dataset_id)
    if dataset is None:
        raise ValueError(f"Dataset {dataset_id} not found")
    return await get_documents_by_dataset(dataset_id)


async def get_questions(dataset_id: UUID) -> list[QuestionORM]:
    dataset = await get_dataset(dataset_id)
    if dataset is None:
        raise ValueError(f"Dataset {dataset_id} not found")
    return await get_questions_by_dataset(dataset_id)


async def ingest_document(
    file: UploadFile, file_url: str, dataset_id: UUID
) -> DocumentORM:
    dataset = await get_dataset(dataset_id)
    if dataset is None:
        raise ValueError(f"Dataset {dataset_id} not found")

    try:
        raw = await file.read()
    except Exception:
        logger.exception(
            "dataset.ingest_document.read_failed",
            extra={
                "event": "dataset.ingest_document.read_failed",
                "dataset_id": str(dataset_id),
            },
        )
        raise

    try:
        blob_url = insert_blob(
            container_name=f"{dataset.name}_{dataset.created_at.strftime('%Y%m%d')}",
            blob_name=file.filename,
            blob_content=raw,
        )
    except Exception:
        logger.exception(
            "dataset.ingest_document.blob_failed",
            extra={
                "event": "dataset.ingest_document.blob_failed",
                "dataset_id": str(dataset_id),
                "blob_name": file.filename or "",
            },
        )
        raise

    document = DocumentORM(
        name=file.filename,
        url=file_url,
        dataset_id=dataset_id,
        blob_url=blob_url,
    )
    try:
        await insert_document(document)
    except Exception:
        logger.exception(
            "dataset.ingest_document.db_failed",
            extra={
                "event": "dataset.ingest_document.db_failed",
                "dataset_id": str(dataset_id),
            },
        )
        raise
    return document


async def ingest_dataset_questions(
    file: UploadFile, dataset_id: UUID
) -> List[QuestionORM]:
    dataset = await get_dataset(dataset_id)
    if dataset is None:
        raise ValueError(f"Dataset {dataset_id} not found")
    try:
        raw = await file.read()
    except Exception as e:
        logger.exception(
            "dataset.ingest_questions.read_failed",
            extra={
                "event": "dataset.ingest_questions.read_failed",
                "dataset_id": str(dataset_id),
            },
        )
        raise ValueError(f"Failed to read upload file: {e}") from e

    try:
        rows = _parse_questions_rows(raw=raw, filename=file.filename)
    except ValueError:
        raise
    except Exception as e:
        logger.exception(
            "dataset.ingest_questions.parse_failed",
            extra={
                "event": "dataset.ingest_questions.parse_failed",
                "dataset_id": str(dataset_id),
            },
        )
        raise ValueError(f"Failed to parse upload file: {e}") from e
    logger.debug(
        "dataset.ingest_questions.parsed",
        extra={
            "event": "dataset.ingest_questions.parsed",
            "dataset_id": str(dataset_id),
            "row_count": len(rows),
            "upload_filename": file.filename,
        },
    )
    dataset_name = dataset.name
    langfuse_client = get_client()
    try:
        langfuse_client.create_dataset(name=dataset_name)
    except Exception as e:
        logger.exception(
            "dataset.ingest_questions.langfuse_dataset_failed",
            extra={
                "event": "dataset.ingest_questions.langfuse_dataset_failed",
                "dataset_id": str(dataset_id),
                "langfuse_dataset": dataset_name,
            },
        )
        raise ValueError(f"Langfuse create_dataset failed: {e}") from e

    questions: list[QuestionORM] = []
    for i, (query, answer, filename) in enumerate(rows, start=2):
        stored_doc_name = f"{dataset.name}_{filename}"
        doc = await find_document_by_name(stored_doc_name, dataset_id)
        if doc is None:
            raise DocumentNotFoundError(i, filename)
        q = QuestionORM(
            query=query,
            answer=answer,
            document_id=doc.id,
            dataset_id=dataset_id,
        )
        try:
            await insert_question(q)
            questions.append(q)
            langfuse_client.create_dataset_item(
                dataset_name=dataset_name,
                input=query,
                expected_output=answer,
                metadata={"id": q.id},
            )
        except Exception as e:
            logger.exception(
                "dataset.ingest_questions.row_failed",
                extra={
                    "event": "dataset.ingest_questions.row_failed",
                    "dataset_id": str(dataset_id),
                    "row_index": i,
                },
            )
            raise ValueError(f"Row {i}: failed to ingest question: {e}") from e

    logger.info(
        "dataset.ingest_questions.complete",
        extra={
            "event": "dataset.ingest_questions.complete",
            "dataset_id": str(dataset_id),
            "inserted": len(questions),
            "langfuse_dataset": dataset_name,
        },
    )
    return questions
