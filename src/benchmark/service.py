from io import BytesIO
from typing import List
from uuid import UUID

from fastapi import UploadFile
from openpyxl import load_workbook

from src.benchmark.repository import (
    find_document_by_name_and_url,
    get_chunks as get_chunks_orm,
    get_embeddings as get_embeddings_orm,
    get_benchmark,
    insert_document,
    insert_chunks as insert_chunks_orm,
    insert_embeddings as insert_embeddings_orm,
    get_scan as get_scan_orm,
    insert_question,
    insert_scan as insert_scan_orm,
)
from src.benchmark.models import DocumentORM, QuestionORM
from src.benchmark.models import ChunkORM, EmbeddingORM, ScanORM
from src.benchmark.schemas import Chunk, Embedding, Scan
from src.infrastructure.blob_storage.blob import insert_blob

_XLSX_COLUMNS = ("query", "answer", "filename", "file_url")


class DocumentNotFoundError(LookupError):
    def __init__(self, row_index: int, name: str, url: str) -> None:
        super().__init__(
            f"Row {row_index}: no document matches filename={name!r} and file_url={url!r}"
        )
        self.row_index = row_index
        self.name = name
        self.url = url


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_xlsx_rows(data: bytes) -> list[tuple[str, str, str, str]]:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []

        headers = [_normalize_header(c) for c in header_row]
        col_index = {h: i for i, h in enumerate(headers) if h}
        for name in _XLSX_COLUMNS:
            if name not in col_index:
                raise ValueError(
                    f"XLSX must contain columns {list(_XLSX_COLUMNS)}; missing {name!r}"
                )

        def cols(r: tuple) -> tuple[str, str, str, str]:
            def at(name: str) -> str:
                i = col_index[name]
                return _cell_str(r[i] if i < len(r) else None)

            return at("query"), at("answer"), at("filename"), at("file_url")

        out: list[tuple[str, str, str, str]] = []
        for row in rows:
            if row is None:
                continue
            q, a, fn, u = cols(row)
            if not (q or a or fn or u):
                continue
            out.append((q, a, fn, u))
        return out
    finally:
        wb.close()


async def ingest_document(
    file: UploadFile, file_url: str, benchmark_id: UUID
) -> DocumentORM:
    benchmark = await get_benchmark(benchmark_id)
    raw = await file.read()
    blob_url = insert_blob(
        container_name=f"{benchmark.name}_{benchmark.created_at.strftime('%Y%m%d')}",
        blob_name=file.filename,
        blob_content=raw,
    )
    document = DocumentORM(
        name=file.filename,
        path=file.filename,
        url=file_url,
        benchmark_id=benchmark_id,
        blob_url=blob_url,
    )
    await insert_document(document)
    return document


async def ingest_benchmark(file: UploadFile, benchmark_id: UUID) -> List[QuestionORM]:
    benchmark = await get_benchmark(benchmark_id)
    if benchmark is None:
        raise ValueError(f"Benchmark {benchmark_id} not found")
    raw = await file.read()
    rows = _parse_xlsx_rows(raw)
    questions: list[QuestionORM] = []
    for i, (query, answer, filename, file_url) in enumerate(rows, start=2):
        doc = await find_document_by_name_and_url(filename, file_url, benchmark_id)
        if doc is None:
            raise DocumentNotFoundError(i, filename, file_url)
        q = QuestionORM(
            query=query,
            answer=answer,
            document_id=doc.id,
            benchmark_id=benchmark_id,
        )
        await insert_question(q)
        questions.append(q)

    return questions


async def get_scan(ocr_id: UUID, document_id: UUID) -> Scan | None:
    obj = await get_scan_orm(ocr_id=ocr_id, document_id=document_id)
    if obj is None:
        return None
    return Scan.model_validate(obj)


async def insert_scan(scan: ScanORM) -> None:
    await insert_scan_orm(scan)


async def get_chunks(chunker_id: UUID, scan_id: UUID) -> list[Chunk]:
    rows = await get_chunks_orm(chunker_id=chunker_id, scan_id=scan_id)
    return [Chunk.model_validate(r) for r in rows]


async def insert_chunks(chunker_id: UUID, chunks: list[ChunkORM]) -> None:
    await insert_chunks_orm(chunker_id=chunker_id, chunks=chunks)


async def get_embeddings(
    embedder_id: UUID, chunker_id: UUID, scan_id: UUID
) -> list[Embedding]:
    rows = await get_embeddings_orm(
        embedder_id=embedder_id, chunker_id=chunker_id, scan_id=scan_id
    )
    return [Embedding.model_validate(r) for r in rows]


async def insert_embeddings(embeddings: list[EmbeddingORM]) -> None:
    await insert_embeddings_orm(embeddings=embeddings)
